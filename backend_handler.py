# backend_handler.py (Adaptado para PySide2, Python 3.8, Win7 32-bit)

# --- Importaciones Estándar ---
import base64
import io
import shutil
import os
import sys
import html
import json
import traceback # Para imprimir errores detallados
from datetime import datetime
import subprocess

# --- Importaciones de Librerías Externas ---
try:
    import qrcode
    QRCODE_AVAILABLE = True
except ImportError:
    print("ADVERTENCIA: Librería 'qrcode' no instalada. Funcionalidad QR deshabilitada.")
    print("Ejecuta: pip install qrcode[pil]")
    QRCODE_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("ADVERTENCIA: Librería 'openpyxl' no instalada. Funcionalidad Excel deshabilitada.")
    print("Ejecuta: pip install openpyxl")
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
                                    Image, PageBreak, BaseDocTemplate, PageTemplate, Frame)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch, cm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.utils import ImageReader # Necesario para QR en PDF
    REPORTLAB_AVAILABLE = True
except ImportError:
    print("¡¡¡ADVERTENCIA!!! ReportLab no está instalado. La exportación a PDF no funcionará.")
    print("Ejecuta: pip install reportlab")
    REPORTLAB_AVAILABLE = False
    # Definir stubs para que el código no falle al cargar
    SimpleDocTemplate = Table = TableStyle = Paragraph = Spacer = Image = PageBreak = None
    BaseDocTemplate = PageTemplate = Frame = None
    getSampleStyleSheet = None
    TA_CENTER = TA_LEFT = TA_RIGHT = 1 # Valores dummy
    cm = 28.346; inch = 72.0 # Valores dummy
    colors = None
    ImageReader = None

# --- Importaciones de PySide2 ---
from PySide2.QtCore import QObject, Slot, Signal, QStandardPaths # <-- CAMBIADO a PySide2
from PySide2.QtWidgets import QFileDialog # <-- CAMBIADO a PySide2 (si se usa para guardar)


# --- Importaciones Locales ---
import database
# La importación de ReportGenerator depende de si ese archivo usa PySide directamente
# Si no usa PySide, la importación es igual. Si usa PySide, debe estar adaptado a PySide2.
try:
    from report_generator import ReportGenerator, REPORTLAB_AVAILABLE
except ImportError:
    print("ADVERTENCIA: No se pudo importar ReportGenerator. Exportación PDF vía clase dedicada deshabilitada.")
    ReportGenerator = None # Definir como None para evitar errores si no se encuentra

    
def resource_path(relative_path):
    """ Obtiene la ruta absoluta al recurso, funciona para dev y PyInstaller """
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(os.path.dirname(__file__)) # Ruta relativa a backend_handler.py
    # ¡Ajusta la ruta relativa si es necesario!
    # Si html_ui está al mismo nivel que backend_handler.py:
    return os.path.join(base_path, relative_path)
    print("--- DEBUG: resource_path function DEFINED in backend_handler.py ---") 

# --- Clase Backend que se expondrá a JavaScript ---
class BackendHandler(QObject):
    # --- Señales (Sintaxis igual en PySide2) ---
    equipment_updated = Signal()
    navigation_requested = Signal(str)
    users_updated = Signal()
    close_application_requested = Signal()
    # print_requested sigue siendo válida si MainWindow la conecta
    print_requested = Signal(list, str)
    test_signal_to_main = Signal(str) # Señal de prueba
    import_excel_finished = Signal(str)      # Envía JSON string del resultado
    equipment_updated = Signal()

    # Nota: La señal show_message fue eliminada previamente.

    def __init__(self, user_role='read_only', user_id=None, parent=None):
        super().__init__(parent)
        self._current_user_role = user_role
        self._current_user_id = user_id
        print(f"BackendHandler (PySide2): Instancia creada. Rol: {self._current_user_role}, ID: {self._current_user_id}")

    # --- Función auxiliar para verificar permisos (sin cambios) ---
    def _check_permission(self, allowed_roles):
        if self._current_user_role not in allowed_roles:
            print(f"Permiso denegado. Rol actual: {self._current_user_role}. Roles permitidos: {allowed_roles}")
            return False
        return True

    # --- Slots (Decorador @Slot igual en PySide2) ---

    @Slot(result=str)
    def get_current_role(self):
        # print("BackendHandler: get_current_role llamado.") # Debug
        return self._current_user_role

    @Slot(result=int)
    def get_current_user_id(self):
        # print(f"BackendHandler: get_current_user_id llamado. Devolviendo: {self._current_user_id}") # Debug
        # Manejar el caso donde user_id podría ser None si el login falla o no lo establece
        return self._current_user_id if self._current_user_id is not None else -1 # Devolver -1 o algún indicador

    # --- Slots para CRUD de Equipos (Lógica sin cambios, dependen de database.py) ---

    @Slot(str, result=str)
    def get_dashboard_data(self, sede_seleccionada=""):
        print(f"BackendHandler: Solicitud get_dashboard_data. Sede seleccionada: '{sede_seleccionada}'")
        try:
            if not self._check_permission(['admin', 'manager', 'read_only']):
                return json.dumps({'error': 'Permiso denegado.'})

            sede_para_filtrar = sede_seleccionada if sede_seleccionada and sede_seleccionada.strip() != "" else None
            print(f"  Backend: Sede para filtrar en DB: {sede_para_filtrar}")

            data = {
                'total_equipos': database.get_total_equipment_count(), # Este es un total general
                'total_usuarios': database.get_total_users_count(),   # Total general
                
                # Campos filtrados por sede:
                'equipos_por_tipo': database.get_count_by_field('tipo_equipo', is_encrypted=False, sede_filtrar=sede_para_filtrar),
                'equipos_por_estatus': database.get_count_by_field('estatus', is_encrypted=False, sede_filtrar=sede_para_filtrar),
                'equipos_en_sede_seleccionada': database.count_equipment_for_sede(sede_para_filtrar),
                'equipos_por_departamento': database.get_count_by_field('departamento', is_encrypted=False, sede_filtrar=sede_para_filtrar),
                
                # Este gráfico no se filtra por sede, ya que ES el gráfico de sedes
                'equipos_por_sede': database.get_count_by_field('sede', is_encrypted=False, sede_filtrar=None),
                
                'mantenimientos_proximos_30d': database.get_proximos_mantenimientos_count(30), # Podría filtrarse por sede también si se desea
                
                'lista_sedes': database.get_distinct_sedes()
            }
            # Debug: Imprimir la cantidad de items para cada categoría después del filtro
            print(f"    Equipos por Tipo (filtrado): {len(data['equipos_por_tipo'])} items")
            print(f"    Equipos por Estatus (filtrado): {len(data['equipos_por_estatus'])} items")
            print(f"    Equipos por Departamento (filtrado): {len(data['equipos_por_departamento'])} items")

            return json.dumps(data)
        except Exception as e:
            print(f"BackendHandler: Error en get_dashboard_data: {e}")
            traceback.print_exc()
            return json.dumps({'error': f'Error generando datos del dashboard: {e}'})

    @Slot(result=list)
    def get_equipment(self):
        # print("BackendHandler: get_equipment llamado.") # Debug
        try:
            equipment_list = database.get_all_equipment()
            # print(f"BackendHandler: Devolviendo {len(equipment_list)} equipos.") # Debug
            return equipment_list
        except Exception as e:
            print(f"BackendHandler: Error en get_equipment: {e}")
            traceback.print_exc()
            return [] # Devolver lista vacía en error
    @Slot(result=str) # Devolver JSON string
    def get_distinct_sedes(self):
        # print("BackendHandler: Solicitud get_distinct_sedes.") # Debug
        try:
            sedes_list = database.get_distinct_sedes() # Llama a la nueva función
            if sedes_list is None: sedes_list = []
            # print(f"  Sedes obtenidas: {sedes_list}") # Debug
            return json.dumps(sedes_list)
        except Exception as e:
            print(f"BackendHandler: Error en get_distinct_sedes: {e}")
            traceback.print_exc()
            return json.dumps([])
        
    @Slot(str, result=bool) # La firma no cambia, pero el manejo del resultado sí.
    def add_equipment(self, data_json_string):
        print(f"BackendHandler: add_equipment llamado con string.")
        response_to_js = False
        try:
            if not self._check_permission(['admin', 'manager']):
                 print("  Permiso denegado.")
                 # Podrías devolver un objeto con más info:
                 # return json.dumps({'success': False, 'error': 'permission_denied'})
                 response_to_js = False # Mantener booleano simple por ahora
            else:
                data_dict = json.loads(data_json_string)
                # Validar serial en el backend antes de llamar a la BD (opcional pero bueno)
                if not data_dict.get('serial') or str(data_dict.get('serial')).strip() == "":
                    print("  Error: Serial es obligatorio y no puede estar vacío (backend).")
                    # Considerar devolver un error específico al JS
                    response_to_js = False
                else:
                    success_db = database.add_new_equipment(data_dict)
                    if success_db: # True si se insertó
                        print("  DB guardó OK.")
                        self.equipment_updated.emit()
                        response_to_js = True
                    else: # False si falló (ej. serial duplicado u otro error de BD)
                        print("  DB falló al guardar (posible serial duplicado o error de validación en BD).")
                        # Aquí es donde el JS necesita saber *por qué* falló.
                        # Por ahora, solo devolvemos False.
                        response_to_js = False

        except json.JSONDecodeError as json_err:
            print(f"  ERROR parseando JSON: {json_err}")
            # return False
            response_to_js = False
        except Exception as e:
            print(f"  ERROR excepción: {e}")
            traceback.print_exc()
            # return False
            response_to_js = False
        finally:
             # Imprimir ANTES de devolver
             print(f"BackendHandler: Devolviendo a JS (add): {response_to_js} (Tipo: {type(response_to_js)})")
             return response_to_js # Devolver el booleano

    @Slot(int, str, str, str, str, result=str) # equipo_id, fecha_str, tipo, desc, meses_str
    def add_maintenance_record(self, equipment_id, fecha_realizado_str, tipo_mantenimiento, descripcion, proximo_en_meses_str):
        print(f"BackendHandler: Solicitud add_maintenance_record para equipo ID: {equipment_id}")
        
        # Permisos (ejemplo: admin o manager pueden registrar mantenimientos)
        if not self._check_permission(['admin', 'manager']):
            return json.dumps({'success': False, 'message': 'Permiso denegado para registrar mantenimiento.'})

        # Convertir proximo_en_meses_str a int o None
        # El string vacío o "0" puede significar que no se programa el próximo
        meses_para_proximo = None
        if proximo_en_meses_str and proximo_en_meses_str.strip():
            try:
                meses_val = int(proximo_en_meses_str.strip())
                if meses_val > 0: # Solo considerar positivo
                    meses_para_proximo = meses_val
                # Si es 0 o negativo, meses_para_proximo queda None (no se programa el siguiente)
            except ValueError:
                return json.dumps({'success': False, 'message': "El valor para 'próximo en meses' debe ser un número."})

        current_user_id = self.get_current_user_id() # Obtener el ID del usuario actual

        success, message = database.add_maintenance(
            equipment_id, 
            fecha_realizado_str, 
            tipo_mantenimiento, 
            descripcion,
            meses_para_proximo, # Ya es int o None
            current_user_id
        )

        if success:
            self.equipment_updated.emit() # Emitir para que la tabla principal se refresque
        
        return json.dumps({'success': success, 'message': message})

    @Slot(int, result=str)
    def get_maintenance_history(self, equipment_id):
        print(f"BackendHandler: Solicitud get_maintenance_history para equipo ID: {equipment_id}")
        # Permisos (ejemplo: cualquiera puede ver el historial si puede ver el equipo)
        if not self._check_permission(['admin', 'manager', 'read_only']):
             return json.dumps({'error': 'Permiso denegado para ver historial.'})

        try:
            history = database.get_maintenance_history_for_equipment(equipment_id)
            return json.dumps({'success': True, 'history': history})
        except Exception as e:
            print(f"Error en get_maintenance_history: {e}")
            traceback.print_exc()
            return json.dumps({'success': False, 'error': f'Error obteniendo historial: {e}'})

    @Slot(int, str, result=bool) # La firma no cambia
    def update_equipment(self, equipment_id, data_json_string):
        print(f"BackendHandler: update_equipment ID={equipment_id} con string.")
        response_to_js = False
        try:
            if not self._check_permission(['admin', 'manager']):
                print("  Permiso denegado.")
                response_to_js = False
            else:
                data_dict = json.loads(data_json_string)
                if not data_dict.get('serial') or str(data_dict.get('serial')).strip() == "":
                    print("  Error: Serial es obligatorio al actualizar (backend).")
                    response_to_js = False
                else:
                    success_db = database.update_existing_equipment(equipment_id, data_dict)
                    if success_db:
                        print("  DB actualizó OK.")
                        self.equipment_updated.emit()
                        response_to_js = True
                    else:
                        print("  DB falló al actualizar (posible serial duplicado o ID no encontrado).")
                        response_to_js = False
                        
        except json.JSONDecodeError as json_err:
            print(f"  ERROR parseando JSON: {json_err}")
            # return False
            response_to_js = False
        except Exception as e:
            print(f"  ERROR excepción: {e}")
            traceback.print_exc()
            # return False
            response_to_js = False
        finally:
             # Imprimir ANTES de devolver
             print(f"BackendHandler: Devolviendo a JS (update): {response_to_js} (Tipo: {type(response_to_js)})")
             return response_to_js # Devolver el booleano

    @Slot(int, result=dict) # Devolver dict
    def delete_equipment(self, equipment_id):
        print(f"BackendHandler: delete_equipment llamado para ID={equipment_id}")
        if not self._check_permission(['admin']):
             return {'success': False, 'message': 'Permiso denegado.'}
        try:
            # Asumiendo que delete_equipment_by_id devuelve True/False
            success = database.delete_equipment_by_id(equipment_id)
            if success:
                self.equipment_updated.emit()
                return {'success': True, 'message': 'Equipo eliminado correctamente.'}
            else:
                 # Podría ser que el ID no existía
                 return {'success': False, 'message': 'Error al eliminar equipo (posiblemente ID no encontrado).'}
        except Exception as e:
            print(f"BackendHandler: Error EXCEPCIÓN en delete_equipment: {e}")
            traceback.print_exc()
            return {'success': False, 'message': f'Error inesperado del servidor: {e}'}

    # --- QR Code Generation (Requiere qrcode y Pillow) ---
    @Slot(int, result=str)
    def get_qr_code_base64(self, equipment_id):
        print(f"BackendHandler: Solicitud get_qr_code_base64 ID={equipment_id}")
        if not QRCODE_AVAILABLE: # Asumiendo que QRCODE_AVAILABLE es un flag global
            print("Error: Librería QR no disponible.")
            return "" # O un Data URI de una imagen de error

        try:
            equipment_details = database.get_equipment_by_id(equipment_id) # Esta función ya desencripta
            if not equipment_details:
                print(f"Error: No se encontró equipo con ID {equipment_id} para generar QR.")
                return ""

            def safe_str(value, default='N/A'): # Helper para formatear
                return html.escape(str(value)) if value is not None and str(value).strip() != "" else default

            # Construir cadena QR con todos los campos relevantes
            # (EXCEPTO historial de mantenimientos)
            qr_content_lines = [
                f"ID Equipo: {equipment_id}",
                f"Tipo: {safe_str(equipment_details.get('tipo_equipo'))}",
                f"Marca: {safe_str(equipment_details.get('marca'))}",
                f"Modelo: {safe_str(equipment_details.get('modelo'))}",
                f"Serial: {safe_str(equipment_details.get('serial'))}",
                f"Asignado a: {safe_str(equipment_details.get('asignado_a'))}",
                f"Departamento: {safe_str(equipment_details.get('departamento'))}",
                f"Sede: {safe_str(equipment_details.get('sede'))}",
                f"Estatus: {safe_str(equipment_details.get('estatus'))}",
                f"Observación: {safe_str(equipment_details.get('observacion'))}",
                f"Registro: {safe_str(equipment_details.get('fecha_registro'))}" # Formatear fecha si es necesario
            ]
            # Filtrar líneas con valor 'N/A' si no se quieren mostrar campos vacíos en el QR
            # qr_content = "\n".join(line for line in qr_content_lines if not line.endswith(": N/A"))
            qr_content = "\n".join(qr_content_lines)


            print(f"  Contenido QR para ID {equipment_id}: {qr_content}")

            qr_img = qrcode.make(qr_content) # make() es más simple para solo contenido de datos
            
            buffered = io.BytesIO()
            qr_img.save(buffered, format="PNG")
            img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
            data_uri = f"data:image/png;base64,{img_str}"
            return data_uri

        except Exception as e:
            print(f"Error generando QR para ID {equipment_id}: {e}")
            traceback.print_exc()
            return ""

    @Slot(result=list)
    def get_users(self):
        print("BackendHandler: get_users llamado.")
        if not self._check_permission(['admin']):
            return []
        try:
            users = database.get_all_users()
            return users if users else []
        except Exception as e:
            print(f"BackendHandler: Error en get_users: {e}")
            traceback.print_exc()
            return []
        
    @Slot(str, str, str, result=str) # Ahora especificamos que devolvemos un string
    def add_user(self, username, password, role):
        response_dict = {} # Diccionario para construir la respuesta
        if not self._check_permission(['admin']):
            response_dict = {'success': False, 'message': 'Permiso denegado.'}
        else:
            try:
                success_db, message_db = database.add_new_user(username, password, role)
                if success_db:
                    self.users_updated.emit()
                response_dict = {'success': success_db, 'message': message_db}
            except Exception as e:
                print(f"BackendHandler: EXCEPCIÓN GRAVE en add_user no manejada por DB: {e}")
                traceback.print_exc()
                response_dict = {'success': False, 'message': f'Error crítico del servidor: {e}'}
        
        json_string_response = json.dumps(response_dict)
        print(f"BackendHandler (add_user): Devolviendo JSON string: {json_string_response}") # Log para ver el string
        return json_string_response

    @Slot(int, str, result=str) # Parámetros: user_id (int), new_role (str). Devuelve: str
    def update_user_role(self, user_id, new_role):
        print(f"BackendHandler: update_user_role llamado para ID {user_id} a rol '{new_role}'.")
        response_dict = {}

        if not self._check_permission(['admin']):
            response_dict = {'success': False, 'message': 'Permiso denegado.'}
        else:
            try:
                # Asumiendo que database.update_user_role devuelve (bool, str)
                success_db, message_db = database.update_user_role(user_id, new_role)
                if success_db:
                    self.users_updated.emit() # Emitir señal si el rol se actualizó
                response_dict = {'success': success_db, 'message': message_db}
            except Exception as e:
                print(f"BackendHandler: Error EXCEPCIÓN en update_user_role: {e}")
                traceback.print_exc()
                response_dict = {'success': False, 'message': f'Error inesperado del servidor: {e}'}
    
        json_string_response = json.dumps(response_dict)
        # ¡ESTE LOG ES CRUCIAL! ¿Qué imprime cuando llamas a editar rol?
        print(f"BackendHandler (update_user_role): Devolviendo JSON string: {json_string_response}") 
        return json_string_response

    @Slot(int, str, result=str) # CAMBIO: result=str
    def reset_password(self, user_id, new_password):
        print(f"BackendHandler: reset_password llamado para ID {user_id}.")
        response_dict = {}

        if not self._check_permission(['admin']):
            response_dict = {'success': False, 'message': 'Permiso denegado.'}
        elif not new_password or len(new_password) < 4:
            response_dict = {'success': False, 'message': "La contraseña debe tener al menos 4 caracteres."}
        else:
            try:
                success_db, message_db = database.reset_user_password(user_id, new_password)
                # No emitimos señal users_updated aquí usualmente
                response_dict = {'success': success_db, 'message': message_db}
            except Exception as e:
                print(f"BackendHandler: Error EXCEPCIÓN en reset_password: {e}")
                traceback.print_exc()
                response_dict = {'success': False, 'message': f'Error inesperado del servidor: {e}'}
                
        json_string_response = json.dumps(response_dict)
        print(f"BackendHandler (reset_password): Devolviendo JSON string: {json_string_response}") # Log útil
        return json_string_response

    @Slot(int, result=str) # CAMBIO: result=str
    def delete_user(self, user_id_to_delete):
        print(f"BackendHandler: Solicitud delete_user para ID {user_id_to_delete}.")
        response_dict = {} # Para construir la respuesta

        if not self._check_permission(['admin']):
            response_dict = {'success': False, 'message': 'Permiso denegado.'}
        elif self._current_user_id is None:
            response_dict = {'success': False, 'message': "Error interno: No se pudo identificar al administrador actual."}
        else:
            try:
                success_db, message_db = database.delete_user_by_id(user_id_to_delete, self._current_user_id)
                if success_db:
                    self.users_updated.emit()
                response_dict = {'success': success_db, 'message': message_db}
            except Exception as e:
                print(f"BackendHandler: Error EXCEPCIÓN en delete_user: {e}")
                traceback.print_exc()
                response_dict = {'success': False, 'message': f'Error inesperado del servidor: {e}'}
        
        json_string_response = json.dumps(response_dict)
        print(f"BackendHandler (delete_user): Devolviendo JSON string: {json_string_response}") # Log útil
        return json_string_response

    # --- Slot para Navegación (sin cambios) ---
    @Slot(str)
    def navigate_to(self, page_name):
        print(f"BackendHandler: Solicitud de navegación a '{page_name}'.")
        self.navigation_requested.emit(page_name)

    # --- Slot para Cerrar Aplicación (sin cambios) ---
    @Slot()
    def request_close_application(self):
        print("BackendHandler: Solicitud de cierre de aplicación recibida desde JS.")
        self.close_application_requested.emit() # Emitir la señal

    # --- Slots para Obtener Valores Distintos (Lógica sin cambios) ---
    @Slot(result=str)
    def get_distinct_marcas(self):
        # print("BackendHandler: Solicitud get_distinct_marcas.")
        try:
            marcas_list = database.get_distinct_field_values('marca')
            if marcas_list is None: marcas_list = []
            # print(f"BackendHandler (get_distinct_marcas): Lista de DB: {marcas_list}")
            return json.dumps(marcas_list)
        except Exception as e:
            print(f"BackendHandler: Error en get_distinct_marcas: {e}")
            traceback.print_exc()
            return json.dumps([])

    @Slot(result=str)
    def get_distinct_modelos(self):
        # print("BackendHandler: Solicitud get_distinct_modelos.")
        try:
            modelos_list = database.get_distinct_field_values('modelo')
            if modelos_list is None: modelos_list = []
            return json.dumps(modelos_list)
        except Exception as e:
            print(f"BackendHandler: Error en get_distinct_modelos: {e}")
            traceback.print_exc()
            return json.dumps([])

    @Slot(result=str)
    def get_distinct_asignados(self):
        # print("BackendHandler: Solicitud get_distinct_asignados.")
        try:
            asignados_list = database.get_distinct_decrypted_field_values('asignado_a')
            if asignados_list is None: asignados_list = []
            return json.dumps(asignados_list)
        except Exception as e:
            print(f"BackendHandler: Error en get_distinct_asignados: {e}")
            traceback.print_exc()
            return json.dumps([])

    @Slot(result=str)
    def get_distinct_departamentos(self):
        # print("BackendHandler: Solicitud get_distinct_departamentos.")
        try:
            deptos_list = database.get_distinct_field_values('departamento')
            if deptos_list is None: deptos_list = []
            return json.dumps(deptos_list)
        except Exception as e:
            print(f"BackendHandler: Error en get_distinct_departamentos: {e}")
            traceback.print_exc()
            return json.dumps([])

    @Slot(result=str)
    def get_distinct_tipos_equipo(self):
        # print("BackendHandler: Solicitud get_distinct_tipos_equipo.")
        try:
            tipos_list = database.get_distinct_field_values('tipo_equipo')
            if tipos_list is None: tipos_list = []
            return json.dumps(tipos_list)
        except Exception as e:
            print(f"BackendHandler: Error en get_distinct_tipos_equipo: {e}")
            traceback.print_exc()
            return json.dumps([])
    

    # --- Slots para Import/Export (Requieren openpyxl, reportlab) ---

    @Slot(result=str) # Ahora devuelve un string JSON
    def download_template_file(self):
        print("BackendHandler: Solicitud download_template_file.")
        response_dict = {'success': False, 'message': 'Error desconocido al descargar plantilla.'} # Default

        try:
            # ... (tu lógica existente para encontrar source_path y download_dir) ...
            source_path = resource_path(os.path.join('html_ui', 'assets', 'template.xlsx')) # Ajusta la ruta si es diferente
            # ... (verificaciones de ruta, os.makedirs) ...

            download_dir = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DownloadLocation)
            if not download_dir or not os.path.isdir(download_dir):
                download_dir = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.HomeLocation)
                if not download_dir or not os.path.isdir(download_dir):
                    download_dir = os.path.abspath(".")
            os.makedirs(download_dir, exist_ok=True)

            suggested_filename = "plantilla_inventario_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
            destination_path = os.path.join(download_dir, suggested_filename)

            shutil.copy2(source_path, destination_path)
            print(f"  Plantilla copiada a {destination_path}")
            friendly_dest_path = destination_path.replace("/", "\\") if os.path.sep == '\\' else destination_path
            
            # --- CRÍTICO: DEVOLVER EL FILEPATH ---
            response_dict = {
                'success': True,
                'message': f'Plantilla descargada en:\n{friendly_dest_path}',
                'filepath': destination_path # <--- AÑADIR ESTA LÍNEA
            }

        except Exception as e:
            print(f"  Error descargando plantilla: {e}")
            traceback.print_exc()
            response_dict = {'success': False, 'message': f'Error al descargar plantilla: {e}'}
        finally:
            json_string_response = json.dumps(response_dict)
            print(f"BackendHandler: Devolviendo a JS (Download Template JSON): {json_string_response}")
            return json_string_response


    @Slot(str) # <-- CAMBIO: Quitado result=...
    def import_excel_data(self, file_base64):
        print("BackendHandler: Solicitud import_excel_data recibida.")
        # Inicializar diccionario de respuesta por defecto
        response_dict = {'success': False, 'message': 'Error desconocido durante importación.', 'inserted': 0, 'db_ignored': 0, 'batch_ignored': 0, 'errors': [], 'db_ignored_details': [], 'batch_ignored_details': []}

        # Usar try...finally para ASEGURAR que la señal se emita
        try:
            # --- Validación inicial ---
            if not OPENPYXL_AVAILABLE:
                 response_dict['message'] = 'Funcionalidad Excel no disponible (falta openpyxl).'
                 # No hacer return aquí, el finally emitirá
            elif not self._check_permission(['admin', 'manager']):
                response_dict['message'] = 'Permiso denegado.'
                 # No hacer return aquí
            elif not file_base64 or not file_base64.startswith('data:application/'):
                 response_dict['message'] = 'Archivo inválido o formato incorrecto.'
                 # No hacer return aquí
            else:
                # --- Si pasa validaciones iniciales, procesar ---
                process_excel = True
                try:
                    header, encoded = file_base64.split(',', 1)
                    excel_bytes = base64.b64decode(encoded)
                except Exception as b64_err:
                     print(f"  Error decodificando Base64: {b64_err}")
                     response_dict['message'] = 'Error al procesar los datos del archivo.'
                     process_excel = False # Marcar para no continuar

                if process_excel:
                    # --- Procesar Excel y llamar a bulk_insert (TU LÓGICA EXISTENTE) ---
                    workbook = openpyxl.load_workbook(filename=io.BytesIO(excel_bytes), data_only=True)
                    sheet = workbook.active
                    header_raw = [cell.value for cell in sheet[1]]
                    header_norm = [str(h).strip() if h is not None else '' for h in header_raw]
                    print(f"  Encabezados leídos: {header_norm}")

                    # Mapeo esperado (TU MAPEADO)
                    expected_header_map = {
                        'Tipo Equipo*': 'tipo_equipo', 'Marca': 'marca', 'Modelo': 'modelo',
                        'Serial*': 'serial', 'Asignado a': 'asignado_a', 'Departamento': 'departamento', 'Sede': 'sede',
                        'Estatus': 'estatus', 'Ultimo Mantenimiento (YYYY-MM-DD)': 'ultimo_mantenimiento',
                        'Proximo Mantenimiento (YYYY-MM-DD)': 'proximo_mantenimiento',
                        'Observacion': 'observacion'
                    }

                    # Validar encabezados y crear mapeo (TU LÓGICA)
                    header_to_index = {}; missing_headers = []
                    for expected_h in expected_header_map.keys():
                        try: actual_idx = header_norm.index(expected_h); header_to_index[expected_h] = actual_idx
                        except ValueError: missing_headers.append(expected_h)

                    if missing_headers:
                        response_dict['message'] = f"Encabezados faltantes o incorrectos: {', '.join(missing_headers)}"
                        response_dict['errors'] = [f"Encabezado faltante: {h}" for h in missing_headers] # Añadir a errores
                    else:
                        data_to_import = []; errors_in_rows = []
                        print(f"  Procesando filas 2 a {sheet.max_row}...")
                        for row_index in range(2, sheet.max_row + 1):
                            # --- TU LÓGICA EXACTA PARA PROCESAR FILAS ---
                            row_values_raw = [cell.value for cell in sheet[row_index]]
                            equipo_dict = {}; has_data = False; row_valid = True
                            for header_excel, db_key in expected_header_map.items():
                                actual_idx = header_to_index[header_excel]
                                value = row_values_raw[actual_idx] if actual_idx < len(row_values_raw) else None
                                processed_value = None
                                if isinstance(value, str): processed_value = value.strip()
                                elif isinstance(value, (int, float)): processed_value = str(value)
                                elif hasattr(value, 'strftime'):
                                    try: processed_value = value.strftime('%Y-%m-%d')
                                    except ValueError: processed_value = None
                                elif value is not None: processed_value = str(value).strip()
                                equipo_dict[db_key] = processed_value if processed_value else None
                                if equipo_dict[db_key] is not None: has_data = True
                            if not has_data: continue
                            if not equipo_dict.get('tipo_equipo'):
                                errors_in_rows.append(f"Fila {row_index}: Tipo Equipo obligatorio está vacío.")
                                row_valid = False
                            # --- (Añade aquí validaciones de SERIAL si las tienes) ---
                            if row_valid:
                                data_to_import.append(equipo_dict)
                            # --- FIN TU LÓGICA PROCESAR FILAS ---

                        print(f"  Se encontraron {len(data_to_import)} filas con datos para intentar importar.")
                        print(f"  Se encontraron {len(errors_in_rows)} errores durante la lectura/validación de filas.")

                        if not data_to_import and not errors_in_rows:
                            response_dict = {'success': True, 'message': 'Archivo Excel vacío o sin filas válidas.', 'inserted': 0, 'db_ignored': 0, 'batch_ignored': 0, 'errors': []}
                        elif not data_to_import and errors_in_rows:
                             response_dict = {'success': False, 'message': 'No se encontraron filas válidas para importar.', 'errors': errors_in_rows}
                        else:
                            # Llamar a bulk_insert
                            print(f"  Llamando a database.bulk_insert_equipment con {len(data_to_import)} registros.")
                            db_result = database.bulk_insert_equipment(data_to_import)

                            # Combinar errores y construir resultado final (TU LÓGICA)
                            all_errors = errors_in_rows + db_result.get('errors', [])
                            total_ignored = db_result.get('db_ignored', 0) + db_result.get('batch_ignored', 0)
                            message = f"Importación: {db_result.get('inserted', 0)} insertados, {total_ignored} ignorados (duplicados), {len(all_errors)} otros errores."

                            if db_result.get('inserted', 0) > 0:
                                 self.equipment_updated.emit()

                            response_dict = {
                                'success': len(all_errors) == 0 and db_result.get('inserted', 0) >= 0, # Ajustado criterio éxito
                                'message': message,
                                'inserted': db_result.get('inserted', 0),
                                'db_ignored': db_result.get('db_ignored', 0),
                                'batch_ignored': db_result.get('batch_ignored', 0),
                                'errors': all_errors,
                                'db_ignored_details': db_result.get('db_ignored_details', []),
                                'batch_ignored_details': db_result.get('batch_ignored_details', [])
                            }
        except Exception as e:
            # Capturar cualquier otra excepción
            print(f"Error GRAL procesando Excel o llamando a bulk: {e}"); traceback.print_exc()
            # Construir un diccionario de error consistente
            response_dict = {
                'success': False,
                'message': f'Error fatal durante la importación: {e}',
                'inserted': 0, 'db_ignored': 0, 'batch_ignored': 0, 'errors': [f'Excepción: {e}'],
                'db_ignored_details': [], 'batch_ignored_details': []
            }
        finally:
            # --- Convertir a JSON y EMITIR SEÑAL ---
            json_string_response = json.dumps(response_dict)
            print(f"BackendHandler: Emitiendo import_excel_finished: {json_string_response}")
            self.import_excel_finished.emit(json_string_response)
            # --- NO HAY RETURN ---

    # --- Slot para Emitir Señal de Impresión (a MainWindow) ---
    @Slot(list, str)
    def request_print_preview(self, filtered_data, print_format):
        """ Recibe datos filtrados y emite señal a MainWindow para imprimir/previsualizar. """
        print(f"BackendHandler: Recibida solicitud request_print_preview ({len(filtered_data)} items, formato: {print_format}).")

        if not isinstance(filtered_data, list):
             print("  Error: Los datos recibidos no son una lista.")
             return
        if print_format not in ['table', 'qr']:
            print(f"  Error: Formato de impresión desconocido: {print_format}")
            return

        try:
            print(f"  Emitiendo señal print_requested...")
            # La señal se conecta en MainWindow, que hará el trabajo de UI (preview/print dialog)
            self.print_requested.emit(filtered_data, print_format)
            print(f"  Señal print_requested emitida.")
        except Exception as e:
            print(f"  Error al emitir señal print_requested: {e}")
            traceback.print_exc()

    # --- Slot para Exportar PDF Directamente (usando ReportGenerator) ---
    @Slot(list, str, result=str) # Recibe: lista de diccionarios de equipos, string de formato
    def export_view_to_pdf(self, equipment_data_list, export_format):
        print(f"BackendHandler: Solicitud export_view_to_pdf. Equipos: {len(equipment_data_list)}, Formato: '{export_format}'")
        response_dict = {'success': False, 'message': 'Error desconocido durante la exportación.'}

        # Validaciones básicas
        if not REPORTLAB_AVAILABLE:
            response_dict['message'] = 'Error: Componente PDF (ReportLab) no está instalado.'
            return json.dumps(response_dict)
        
        # ReportGenerator podría no estar disponible si su importación falló
        if ReportGenerator is None:
             response_dict['message'] = 'Error: Generador de reportes no inicializado.'
             return json.dumps(response_dict)

        if not self._check_permission(['admin', 'manager', 'read_only']): # Asumiendo que todos pueden exportar
            response_dict['message'] = 'Permiso denegado para exportar.'
            return json.dumps(response_dict)

        if not isinstance(equipment_data_list, list):
            response_dict['message'] = 'Error: Datos de entrada inválidos (no es una lista).'
            return json.dumps(response_dict)

        if not equipment_data_list:
            response_dict['message'] = 'No hay datos para exportar.'
            # Devolver success: True pero con filepath: None podría ser una opción
            # para que el JS no muestre un error, sino un mensaje informativo.
            # Por ahora, lo dejamos como fallo para que JS muestre el mensaje.
            return json.dumps(response_dict)

        if export_format not in ['table', 'qr']:
            response_dict['message'] = f"Formato de exportación no soportado: '{export_format}'."
            return json.dumps(response_dict)

        if export_format == 'qr' and not QRCODE_AVAILABLE:
            response_dict['message'] = 'Error: Componente QR (qrcode) no está instalado para exportar códigos QR.'
            return json.dumps(response_dict)

        try:
            generator = ReportGenerator() # Crear instancia del generador
            
            print(f"  BackendHandler: Delegando a ReportGenerator.generate_{export_format}_report...")
            
            if export_format == 'table':
                # generate_table_report espera la lista de diccionarios de equipos
                result_from_generator = generator.generate_table_report(equipment_data_list)
            elif export_format == 'qr':
                # generate_qr_report también espera la lista de diccionarios de equipos
                result_from_generator = generator.generate_qr_report(equipment_data_list)
            else: # No debería llegar aquí por la validación anterior
                result_from_generator = {'success': False, 'message': 'Formato de exportación interno no reconocido.'}

            # El generador debería devolver un dict con {'success': bool, 'message': str, 'filepath': str (opcional)}
            if isinstance(result_from_generator, dict):
                response_dict = result_from_generator
            else:
                print(f"  ADVERTENCIA: ReportGenerator no devolvió un diccionario. Respuesta: {result_from_generator}")
                response_dict = {'success': False, 'message': 'Error interno del servidor al procesar el reporte.'}
        
        except Exception as e_export:
            print(f"!!! EXCEPCIÓN GRAVE en export_view_to_pdf: {e_export}")
            traceback.print_exc()
            response_dict = {'success': False, 'message': f"Error crítico del servidor durante la exportación: {e_export}"}
        
        json_string_response = json.dumps(response_dict)
        print(f"BackendHandler (export_view_to_pdf): Devolviendo JSON: {json_string_response}")
        return json_string_response
        
    @Slot(str, result=str) # Recibe JSON string de lista de IDs
    def bulk_delete_equipment(self, ids_json_string):
        print(f"BackendHandler: Solicitud bulk_delete_equipment con IDs: {ids_json_string}")
        if not self._check_permission(['admin']):
            return json.dumps({'success': False, 'message': 'Permiso denegado para eliminar equipos.'})
        
        try:
            ids_to_delete = json.loads(ids_json_string)
            if not isinstance(ids_to_delete, list) or not ids_to_delete:
                return json.dumps({'success': False, 'message': 'IDs de equipos inválidos o vacíos.'})
            
            deleted_count = database.delete_multiple_equipment_by_ids(ids_to_delete, self._current_user_id) # Nueva función en database
            
            if deleted_count > 0:
                self.equipment_updated.emit(); # Notificar al frontend
                return json.dumps({'success': True, 'message': f'{deleted_count} equipo(s) eliminado(s) correctamente.'})
            else:
                return json.dumps({'success': False, 'message': 'Ningún equipo fue eliminado (posiblemente no encontrado o sin permiso para eliminar el último admin).'})
        except json.JSONDecodeError:
            return json.dumps({'success': False, 'message': 'Formato de IDs JSON inválido.'})
        except Exception as e:
            print(f"Error en bulk_delete_equipment: {e}"); traceback.print_exc()
            return json.dumps({'success': False, 'message': f'Error del servidor al eliminar equipos: {e}'})

    @Slot(str, str, result=str) # Recibe IDs JSON, y Updates JSON
    def update_multiple_equipment(self, ids_json_string, updates_json_string):
        print(f"BackendHandler: Solicitud update_multiple_equipment para IDs: {ids_json_string}, Updates: {updates_json_string}")
        if not self._check_permission(['admin', 'manager']):
            return json.dumps({'success': False, 'message': 'Permiso denegado para editar equipos.'})

        try:
            ids_to_update = json.loads(ids_json_string)
            updates_data = json.loads(updates_json_string)

            if not isinstance(ids_to_update, list) or not ids_to_update:
                return json.dumps({'success': False, 'message': 'IDs de equipos inválidos o vacíos.'})
            if not isinstance(updates_data, dict) or not updates_data:
                return json.dumps({'success': False, 'message': 'Datos de actualización inválidos o vacíos.'})
            
            # Llamar a la función de la base de datos para la actualización masiva
            updated_count, errors = database.update_multiple_equipment_fields(ids_to_update, updates_data) # Nueva función en database

            if updated_count > 0:
                self.equipment_updated.emit()
                return json.dumps({'success': True, 'message': f'{updated_count} equipo(s) actualizado(s) correctamente.'})
            elif errors:
                return json.dumps({'success': False, 'message': f'Error al actualizar algunos equipos: {errors}'})
            else:
                return json.dumps({'success': False, 'message': 'Ningún equipo fue actualizado (posiblemente no encontrado o no hubo cambios válidos).'})
        except json.JSONDecodeError:
            return json.dumps({'success': False, 'message': 'Formato JSON inválido para IDs o actualizaciones.'})
        except Exception as e:
            print(f"Error en update_multiple_equipment: {e}"); traceback.print_exc()
            return json.dumps({'success': False, 'message': f'Error del servidor al actualizar equipos: {e}'})    
    @Slot(str)
    def open_file(self, filepath):
        """Intenta abrir un archivo usando la aplicación predeterminada del SO."""
        print(f"BackendHandler: Solicitud para abrir archivo: {filepath}")
        if not filepath or not os.path.exists(filepath):
            print(f"  Error: El archivo no existe o la ruta es inválida: {filepath}")
            # Opcional: Emitir señal de error a JS si quieres notificar al usuario
            # self.show_message.emit('error', f"No se pudo encontrar el archivo:\n{filepath}")
            return

        try:
            if sys.platform == "win32":
                os.startfile(os.path.normpath(filepath)) # os.startfile es para Windows
                print(f"  Intentando abrir con os.startfile...")
            elif sys.platform == "darwin": # macOS
                subprocess.run(['open', filepath], check=True)
                print(f"  Intentando abrir con 'open' (macOS)...")
            else: # Linux y otros Unix-like
                subprocess.run(['xdg-open', filepath], check=True)
                print(f"  Intentando abrir con 'xdg-open' (Linux)...")
            print(f"  Comando para abrir archivo ejecutado.")
        except FileNotFoundError:
             print(f"  Error: No se encontró el comando para abrir archivos ('open' o 'xdg-open').")
             # self.show_message.emit('error', "No se encontró la aplicación para abrir el archivo.")
        except subprocess.CalledProcessError as e:
             print(f"  Error al ejecutar comando para abrir archivo: {e}")
             # self.show_message.emit('error', f"Error al intentar abrir el archivo:\n{e}")
        except Exception as e:
            print(f"!!! Error inesperado al intentar abrir archivo: {e}")
            traceback.print_exc()


# --- Bloque de prueba (Opcional) ---
if __name__ == '__main__':
     print("Probando BackendHandler (requiere entorno Qt para señales)...")
     # Para probar realmente las señales, necesitarías una QApplication
     # app = QApplication([]) # Descomentar si necesitas probar señales
     handler_admin = BackendHandler(user_role='admin', user_id=1)
     handler_readonly = BackendHandler(user_role='read_only', user_id=2)

     print("\n--- Probando Permisos ---")
     print(f"Admin puede borrar? {handler_admin._check_permission(['admin'])}")
     print(f"Read-only puede borrar? {handler_readonly._check_permission(['admin'])}")
     print(f"Read-only puede ver? {handler_readonly._check_permission(['admin', 'manager', 'read_only'])}")

     print("\n--- Probando getters ---")
     print(f"Rol admin: {handler_admin.get_current_role()}")
     print(f"ID admin: {handler_admin.get_current_user_id()}")
     print(f"Rol read-only: {handler_readonly.get_current_role()}")
     print(f"ID read-only: {handler_readonly.get_current_user_id()}")

     # Aquí podrías añadir llamadas a otros métodos si tienes una BD de prueba
     # print("\n--- Probando get_equipment ---")
     # equipos = handler_admin.get_equipment()
     # print(f"Equipos obtenidos: {len(equipos)}")

     print("\nFin prueba BackendHandler.")
     # if 'app' in locals(): sys.exit(app.exec_()) # Descomentar si usaste QApplication